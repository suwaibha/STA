from faker import Faker
from openpyxl import Workbook

# Create workbook
wb = Workbook()
ws = wb.active

# Create Faker object
fake_data = Faker()

# Generate 10 fake records
for i in range(1, 11):
    ws.cell(row=i, column=1).value = fake_data.name()
    ws.cell(row=i, column=2).value = fake_data.email()

# Save the Excel file
wb.save("demo.xlsx")

print("Excel file 'demo.xlsx' created successfully.")