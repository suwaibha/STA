from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import Workbook, load_workbook

import os
import time

driver = webdriver.Chrome()

driver.maximize_window()

file_path = os.path.abspath("export.html")

driver.get("file:///" + file_path.replace("\\","/"))

wait = WebDriverWait(driver,10)

# ---------------- Apply Filter ----------------

wait.until(
    EC.element_to_be_clickable((By.ID,"filterBtn"))
).click()

alert = wait.until(EC.alert_is_present())
print(alert.text)
alert.accept()

# ---------------- Read UI Data ----------------

rows = driver.find_elements(By.XPATH,"//*[@id='studentTable']/tbody/tr")

uiData=[]

for row in rows:

    cols=row.find_elements(By.TAG_NAME,"td")

    uiData.append([cols[0].text,cols[1].text])

print("\nRows in UI :",len(uiData))

# ---------------- Export ----------------

driver.find_element(By.ID,"exportBtn").click()

alert = wait.until(EC.alert_is_present())
print(alert.text)
alert.accept()

# ---------------- Create Excel ----------------

wb=Workbook()

ws=wb.active

ws.append(["ID","Name"])

for row in uiData:

    ws.append(row)

excelFile="Student_Report.xlsx"

wb.save(excelFile)

print("\nExcel Generated Successfully")

# ---------------- Verify Download ----------------

if os.path.exists(excelFile):

    print("PASS : Excel File Exists")

else:

    print("FAIL : Excel File Missing")

# ---------------- Read Excel ----------------

wb=load_workbook(excelFile)

ws=wb.active

excelRows=ws.max_row-1

print("Rows in Excel :",excelRows)

# ---------------- Validation ----------------

assert excelRows==len(uiData)

print("\nPASS : Excel data matches UI")

input("\nPress Enter to Close Browser...")

driver.quit()